import shutil
import zipfile
import shell
from openpyxl import load_workbook
from pypdf import PdfReader
import xlrd
import csv
import time
import os.path
import requests
from selenium import webdriver
from selene import browser
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from win32com.shell import shell, shellcon


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_xlsx():
    PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
    workbook_path = os.path.join(PROJECT_ROOT_PATH, 'resources', 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(workbook_path)
    sheet = workbook.active
    value = sheet.cell(row=3, column=2).value
    assert value == 'Mara'


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_xlsx_2():
    print(end='\n')
    print('')
    PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
    workbook_path = os.path.join(PROJECT_ROOT_PATH, 'resources', 'file_example_XLS_10.xls')
    book = xlrd.open_workbook(workbook_path)
    print(f'Количество листов {book.nsheets}')
    print(f'Имена листов {book.sheet_names()}')
    sheet = book.sheet_by_index(0)
    print(f'Количество столбцов {sheet.ncols}')
    print(f'Количество строк {sheet.nrows}')
    print(f'Пересечение строки 9 и столбца 1 = {sheet.cell_value(rowx=0, colx=1)}')
    # печать всех строк по очереди
    for rx in range(sheet.nrows):
        print(sheet.row(rx))

    assert book.nsheets == 1
    assert sheet.ncols == 8
    assert sheet.cell_value(rowx=0, colx=1) == 'First Name'


# TODO оформить в тест, добавить ассерты и использовать универсальный путь

def test_pdf():
    PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
    pdf_path = os.path.join(PROJECT_ROOT_PATH, 'resources', 'docs-pytest-org-en-latest.pdf')
    reader = PdfReader(pdf_path)
    number_of_pages = len(reader.pages)
    page = reader.pages[0]
    text = page.extract_text()
    print(end='\n')
    print('')
    print(page)
    print(number_of_pages)
    print(text)
    assert number_of_pages == 412


def test_csv_file():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
    csv_path = os.path.join(PROJECT_ROOT_PATH, 'resources', 'eggs.csv')
    with open(csv_path, 'w') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=',')
        csvwriter.writerow(['Anna', 'Pavel', 'Peter'])
        csvwriter.writerow(['Alex', 'Serj', 'Yana'])

    with open(csv_path) as csvfile:
        csvreader = csv.reader(csvfile)
        save = []
        for row in csvreader:
            save.append(row)
            print(row)
    assert save[0] == ['Anna', 'Pavel', 'Peter']


def test_downloaded_file_size():
    # TODO сохранять и читать из tmp, использовать универсальный путь
    url = 'https://demoqa.com/images/Toolsqa.jpg'
    PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
    tmp_file = os.path.join(PROJECT_ROOT_PATH, 'tmp', 'Toolsqa.jpg')
    r = requests.get(url)
    with open(tmp_file, 'wb') as file:
        file.write(r.content)
    size = os.path.getsize(tmp_file)
    assert size == 7232


def test_download_big_file_size():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь к tmp
    PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
    download_folder = shell.SHGetKnownFolderPath(shellcon.FOLDERID_Downloads)
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": 'download_folder',
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )

    browser.config.driver = driver
    browser.open("https://demoqa.com/upload-download")
    browser.element('[id = downloadButton]').click()
    time.sleep(10)
    filename_old = os.path.join(download_folder, 'sampleFile.jpeg')
    filename_new = os.path.join(PROJECT_ROOT_PATH, 'tmp', 'sampleFile.jpeg')
    new_location = shutil.move(filename_old, filename_new)
    downloaded_file_size = os.path.getsize(new_location)
    assert downloaded_file_size == 4096


def test_add_files_to_zip():
    PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
    files_to_zip = os.path.join(PROJECT_ROOT_PATH, 'resources')
    files_to_zip_list = os.listdir(files_to_zip)
    with zipfile.ZipFile('test.zip', mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for file in files_to_zip_list:
            add_file = os.path.join(files_to_zip, file)
            zf.write(add_file)

    file_dict = {'docs-pytest-org-en-latest.pdf': 1739253,
                 'eggs.csv': 36,
                 'file_example_XLSX_50.xlsx': 7360,
                 'file_example_XLS_10.xls': 8704,
                 'hello.zip': 128}
    print('')
    with zipfile.ZipFile('test.zip', mode='a') as zf:
        for file in zf.infolist():
            name = os.path.basename(file.filename)
            print(f"{name}, {file.file_size}")
            assert file.file_size == file_dict.get(name)
